from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import openpyxl
from webdriver_manager.chrome import ChromeDriverManager

print("Iniciando nosso robô...\n")
arq = open("resultado.txt", "w")

dominios = []
# Lendo do Excel
workbook = openpyxl.load_workbook('dominios.xlsx')
sheet = workbook.active

for linha in range(1, sheet.max_row + 1):
    dominios.append(sheet.cell(row=linha, column=1).value)

driver = webdriver.Chrome(ChromeDriverManager().install())
driver.get("https://registro.br/")

for dominio in dominios:
    pesquisa = driver.find_element("id", "is-avail-field")
    pesquisa.clear()  # Limpando a barra de pesquisa
    pesquisa.send_keys(dominio)
    pesquisa.send_keys(Keys.RETURN)
    time.sleep(2)
    resultados = driver.find_elements("tag name", "strong")
    texto = "Domínio %s %s\n" % (dominio, resultados[4].text)
    arq.write(texto)

arq.close()
driver.close()
